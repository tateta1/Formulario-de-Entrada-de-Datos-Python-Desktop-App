import os.path
import re
import tkinter as tk
from ast import pattern
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import re
import  os

from openpyxl.reader.excel import load_workbook

nombre_archivo = 'datos.xlsx'
#Comprobar si el archivo ya existe
if os.path.exists(nombre_archivo):
        wb = load_workbook(nombre_archivo)
        ws = wb.active

# Configuracion de Exel
wb = Workbook()
ws = wb.active
ws.append(['Nombre', "Edad", "Email", "Telefono","Direccion",])

def guardar_datos():
    nombre = entry_nombre.get()
    edad = entry_edad.get()
    email = entry_email.get()
    telefono = entry_telefono.get()
    direccion = entry_direccion.get()

    if not nombre or not edad or not email or not telefono or not direccion:
        messagebox.showwarning(title="Advertencia", message="Todos los campos son obligatorios")
        return

    try:
        edad = int (edad)
        telefono = int(telefono)
    except ValueError:
        messagebox.showwarning(title="Advertencia", message="Edad y Telefono deben ser números")
        return

    #Validar Formato de email
    if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
         messagebox.showwarning(title="Advertencia", message="El correo eletronico no es valido")
         return

    ws.append([nombre, edad, email, telefono, direccion])
    wb.save('nombre_archivo.xlsx')
    messagebox.showinfo(title="Información", message="Datos guardados con éxito")

    #Esto seria para cuando le des guadar los datos automaticamente  se borren
    entry_nombre.delete(first=0, last=tk.END)
    entry_edad.delete(first=0, last= tk.END)
    entry_email.delete(first=0, last=tk.END)
    entry_telefono.delete(first=0, last= tk.END)
    entry_direccion.delete(first=0, last= tk.END)


root = tk.Tk()
root.title("Formulario de Entrada de Datos")
root.configure(bg='#4b6587')
label_style = {"bg": '#4b6587', "fg": "white"}
entry_style = {"bg": '#D3D3D3', "fg": "black"}

label_nombre = tk.Label(root, text="Nombre", **label_style)
label_nombre.grid(row=0, column=0, padx=10, pady=5)
entry_nombre = tk.Entry(root, **entry_style)
entry_nombre.grid(row=0, column=1, padx=10, pady=5)

label_edad = tk.Label(root, text="Edad", **label_style)
label_edad.grid(row=1, column=0, padx=10, pady=5)
entry_edad = tk.Entry(root, **entry_style)
entry_edad.grid(row=1, column=1, padx=10, pady=5)


label_email = tk.Label(root, text="Email", **label_style)
label_email.grid(row=2, column=0, padx=10, pady=5)
entry_email = tk.Entry(root, **entry_style)
entry_email.grid(row=2, column=1, padx=10, pady=5)


label_telefono = tk.Label(root, text="Telefono", **label_style)
label_telefono.grid(row=3, column=0, padx=10, pady=5)
entry_telefono = tk.Entry(root, **entry_style)
entry_telefono.grid(row=3, column=1, padx=10, pady=5)

label_direccion = tk.Label(root, text="Direccion", **label_style)
label_direccion.grid(row=4, column=0, padx=10, pady=5)
entry_direccion = tk.Entry(root, **entry_style)
entry_direccion.grid(row=4, column=1, padx=10, pady=5)



boton_guardar = tk.Button(root, text="Guardar",command=guardar_datos,
                          bg='#6D8299', fg='white', width=30)
boton_guardar.grid(row=5, column=0, columnspan=2, padx=10, pady=10)
root.mainloop()